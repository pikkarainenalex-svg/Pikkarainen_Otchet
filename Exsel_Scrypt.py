import openpyxl
from openpyxl import Workbook
import os

wb = Workbook()
ws = wb.active
ws.title = "Haulhwr"

ws['A1'] = "Заголовок 1"
ws['B1'] = "Заголовок 2"
ws['C1'] = "Заголовок 3"

data = [
    ["Иван", "Иванович", 24],
    ["Александр", "Владимирович", 32],
    ["Максим", "Василович", 44]
]

for row in data:
    ws.append(row)

script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, 'example.xlsx')

wb.save(file_path)

print(f"Файл сохранен: {file_path}")
print("Содержимое файла example.xlsx:")
for row in ws.iter_rows(values_only=True):
    print("\t".join(str(cell) if cell is not None else "" for cell in row))








